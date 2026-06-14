#!/usr/bin/env python3
import sys
import subprocess
import json

FILE = sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\ay036\IdeaProjects\EmployeeManagement\Party_wise_Sales_Summary.xlsx"

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed, attempting to install...", file=sys.stderr)
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook

try:
    wb = load_workbook(FILE, read_only=True, data_only=True)
except Exception as e:
    print(f"ERROR: Could not open workbook {FILE}: {e}", file=sys.stderr)
    sys.exit(2)

sheets = wb.sheetnames
print(json.dumps({"sheets": sheets}))

# Read first sheet
ws = wb[sheets[0]]
rows = ws.iter_rows(values_only=True)

preview = []
for i, row in enumerate(rows):
    if i >= 11:
        break
    # convert to strings safely
    preview.append([None if v is None else str(v) for v in row])

print(json.dumps({"preview_rows": preview}, ensure_ascii=False))

